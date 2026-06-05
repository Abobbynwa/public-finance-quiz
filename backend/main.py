import os
import json
import uuid
import time
from datetime import datetime
from typing import Optional

import firebase_admin
from firebase_admin import credentials, firestore
from fastapi import BackgroundTasks, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook

ADMIN_TOKEN = os.getenv("EDUCBT_ADMIN_TOKEN", "change-this-token")
ALLOWED_ORIGINS = [x.strip() for x in os.getenv("EDUCBT_ALLOWED_ORIGINS", "*").split(",") if x.strip()]
BATCH_SIZE = int(os.getenv("EDUCBT_IMPORT_BATCH_SIZE", "100"))
BATCH_DELAY = float(os.getenv("EDUCBT_IMPORT_BATCH_DELAY", "1.2"))

app = FastAPI(title="EduCBT Backend", version="1.2.0")
app.add_middleware(CORSMiddleware, allow_origins=ALLOWED_ORIGINS, allow_credentials=False, allow_methods=["*"], allow_headers=["*"])
JOBS = {}

VALID_SUBJECTS = {"Economics", "Government", "Commerce", "Accounting", "English Language", "Mathematics", "Biology", "Chemistry", "Physics", "Agricultural Science", "Civic Education", "CRS", "Basic Science", "Basic Technology", "Social Studies", "Business Studies", "Computer Studies", "Home Economics", "PHE"}
VALID_CLASSES = {"JSS", "JSS1", "JSS2", "JSS3", "SS", "SS1", "SS2", "SS3", "GENERAL", ""}
VALID_ANSWERS = {"A", "B", "C", "D"}


def init_firebase():
    if firebase_admin._apps:
        return firestore.client()
    service_json = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON")
    service_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    if service_json:
        cred = credentials.Certificate(json.loads(service_json))
    elif service_path:
        cred = credentials.Certificate(service_path)
    else:
        raise RuntimeError("Firebase credentials missing")
    firebase_admin.initialize_app(cred)
    return firestore.client()


def verify_admin(token: Optional[str]):
    if not token or token != ADMIN_TOKEN:
        raise HTTPException(status_code=401, detail="Unauthorized admin request")


def norm(value):
    return str(value or "").strip().lower().replace("_", " ")


def get_value(row, headers, names):
    wanted = [norm(x) for x in names]
    for index, header in enumerate(headers):
        if norm(header) in wanted:
            return row[index]
    return ""


def commit_batch(batch, sleep_after=True):
    batch.commit()
    if sleep_after and BATCH_DELAY > 0:
        time.sleep(BATCH_DELAY)


class Settings(BaseModel):
    schoolName: str = "EduCBT Portal"
    logoUrl: str = ""
    motto: str = "Knowledge and Excellence"
    schoolCode: str = "EDU"
    term: str = "First Term"
    session: str = "2025/2026"
    examType: str = "Quiz"
    questionsPerExam: int = 20
    timeMinutes: int = 15
    allowRetake: str = "yes"
    showCorrection: str = "yes"


@app.get("/")
def root():
    return {"ok": True, "service": "EduCBT Backend", "docs": "/docs"}


@app.get("/health")
def health():
    return {"ok": True, "service": "EduCBT Backend", "time": datetime.utcnow().isoformat()}


@app.post("/admin/settings")
def save_settings(settings: Settings, x_admin_token: Optional[str] = Header(default=None)):
    verify_admin(x_admin_token)
    init_firebase().collection("settings").document("exam").set(settings.dict(), merge=True)
    return {"ok": True, "message": "Settings saved"}


@app.get("/admin/settings")
def get_settings(x_admin_token: Optional[str] = Header(default=None)):
    verify_admin(x_admin_token)
    doc = init_firebase().collection("settings").document("exam").get()
    return doc.to_dict() if doc.exists else Settings().dict()


def import_worker(job_id: str, temp_path: str, target_class: str, mode: str):
    try:
        JOBS[job_id].update({"status": "processing", "message": "Reading Excel file"})
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Empty Excel file")
        headers = [str(x or "").strip() for x in rows[0]]
        parsed = []
        for row_number, row in enumerate(rows[1:], start=2):
            subject = str(get_value(row, headers, ["Subject"]) or "").strip()
            if mode != "ALL" and not subject:
                subject = mode
            raw_class = str(get_value(row, headers, ["Class", "Class Level", "Level"]) or "").strip().upper()
            fallback_class = str(target_class or "").strip().upper()
            class_level = raw_class or fallback_class
            if class_level == "GENERAL":
                class_level = ""
            question = str(get_value(row, headers, ["Question", "q"]) or "").strip()
            options = [
                str(get_value(row, headers, ["A", "Option A"]) or "").strip(),
                str(get_value(row, headers, ["B", "Option B"]) or "").strip(),
                str(get_value(row, headers, ["C", "Option C"]) or "").strip(),
                str(get_value(row, headers, ["D", "Option D"]) or "").strip(),
            ]
            answer = str(get_value(row, headers, ["Correct Answer", "Answer"]) or "").strip().upper()
            if not subject and not question:
                continue
            if subject not in VALID_SUBJECTS:
                raise ValueError(f"Invalid subject on row {row_number}: {subject}")
            if class_level not in VALID_CLASSES:
                raise ValueError(f"Invalid class on row {row_number}: {class_level}")
            if not question or any(not option for option in options):
                raise ValueError(f"Missing question/options on row {row_number}")
            if answer not in VALID_ANSWERS:
                raise ValueError(f"Invalid answer on row {row_number}: {answer}")
            parsed.append({"classLevel": class_level, "subject": subject, "q": question, "options": options, "answer": answer, "active": True, "createdAt": firestore.SERVER_TIMESTAMP})
        if not parsed:
            raise ValueError("No valid questions found")

        db = init_firebase()
        pairs = {(item["classLevel"], item["subject"]) for item in parsed}
        JOBS[job_id].update({"total": len(parsed), "groups": len(pairs), "message": "Deactivating old matching questions", "batchSize": BATCH_SIZE})

        batch = db.batch(); ops = 0; deactivated = 0
        for doc in db.collection("quizQuestions").stream():
            data = doc.to_dict()
            pair = (str(data.get("classLevel", "")).upper(), data.get("subject", ""))
            if pair in pairs:
                batch.update(doc.reference, {"active": False})
                ops += 1; deactivated += 1
                if ops >= BATCH_SIZE:
                    commit_batch(batch); batch = db.batch(); ops = 0
                    JOBS[job_id].update({"deactivated": deactivated, "message": f"Deactivated {deactivated} old questions"})
        if ops:
            commit_batch(batch); batch = db.batch(); ops = 0

        imported = 0
        JOBS[job_id].update({"message": "Importing new questions slowly to avoid quota limit", "deactivated": deactivated})
        for item in parsed:
            batch.set(db.collection("quizQuestions").document(), item)
            ops += 1; imported += 1
            if ops >= BATCH_SIZE:
                commit_batch(batch); batch = db.batch(); ops = 0
                JOBS[job_id].update({"imported": imported, "message": f"Imported {imported}/{len(parsed)} questions"})
        if ops:
            commit_batch(batch, sleep_after=False)
        JOBS[job_id].update({"status": "done", "imported": imported, "message": f"Imported {imported} questions safely"})
    except Exception as exc:
        JOBS[job_id].update({"status": "error", "message": str(exc)})
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


@app.post("/admin/questions/import")
async def import_questions(background_tasks: BackgroundTasks, target_class: str = Form(""), mode: str = Form("ALL"), file: UploadFile = File(...), x_admin_token: Optional[str] = Header(default=None)):
    verify_admin(x_admin_token)
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Upload an Excel .xlsx file")
    job_id = str(uuid.uuid4())
    temp_path = f"/tmp/educbt-{job_id}.xlsx"
    with open(temp_path, "wb") as f:
        f.write(await file.read())
    JOBS[job_id] = {"job_id": job_id, "status": "queued", "message": "Upload received", "imported": 0, "total": 0, "createdAt": datetime.utcnow().isoformat()}
    background_tasks.add_task(import_worker, job_id, temp_path, target_class, mode)
    return {"ok": True, "queued": True, "job_id": job_id, "message": "Import started. Keep this page open and wait for completion."}


@app.get("/admin/jobs/{job_id}")
def get_job(job_id: str, x_admin_token: Optional[str] = Header(default=None)):
    verify_admin(x_admin_token)
    if job_id not in JOBS:
        raise HTTPException(status_code=404, detail="Job not found")
    return JOBS[job_id]


@app.get("/admin/results")
def list_results(className: Optional[str] = None, subject: Optional[str] = None, term: Optional[str] = None, session: Optional[str] = None, x_admin_token: Optional[str] = Header(default=None)):
    verify_admin(x_admin_token)
    query = init_firebase().collection("quizResults")
    if className:
        query = query.where("className", "==", className)
    if subject:
        query = query.where("subject", "==", subject)
    if term:
        query = query.where("term", "==", term)
    if session:
        query = query.where("session", "==", session)
    results = []
    for doc in query.limit(500).stream():
        item = doc.to_dict(); item["id"] = doc.id; results.append(item)
    return {"ok": True, "count": len(results), "results": results}
